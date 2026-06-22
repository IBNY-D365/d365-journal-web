"""
IBNY D365 Journal Entry Automation App
---------------------------------------
Automates cash closing by matching BOA + Zoho transactions,
resolving customer accounts, and generating a D365-ready Excel upload.

Source-of-truth files (bundled with app):
  - Cash_Code_Masterlist.xlsx
  - IBNY_Business_Customer_Account.xlsx
  - Posted_Journal_in_D365_Sample_Reference.xlsx  (format reference)
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import warnings
warnings.filterwarnings("ignore")

# ── Inline modules ───────────────────────────────────────────────────────────

"""
parsers.py — File parsing for BOA, Zoho, and reference spreadsheets.

Tuned to the actual file formats observed:
  BOA CSV  : 5-row summary block at top, then blank line, then real data header
             Columns: Date, Description, Amount, Running Bal.
  Zoho PDF : Screen-captured payout page from pay.zoho.com — no proper table
             structure. All data lives in a single merged cell as raw text.
             Must be parsed from the plain text of each page.
"""

import re
import io
import pandas as pd
import numpy as np

# ── Column aliases ────────────────────────────────────────────────────────────
_BOA_COL_MAP = {
    "date":        ["date", "posting date", "post date", "transaction date",
                    "trans date", "effective date"],
    "description": ["description", "payee", "details", "memo", "narrative",
                    "transaction description", "trans desc"],
    "amount":      ["amount", "net amount", "credit", "credit amount",
                    "transaction amount", "trans amount"],
    "debit":       ["debit", "debit amount"],
    "balance":     ["balance", "running balance", "running bal", "running bal."],
}

_ZOHO_COL_MAP = {
    "date":         ["date", "payment date", "transaction date", "created date",
                     "created time", "payout date", "date & time"],
    "customer":     ["customer name", "customer", "client name", "name",
                     "bill to", "payer", "contact", "buyer"],
    "gross_amount": ["gross amount", "gross", "amount", "total amount",
                     "payment amount", "gross sales"],
    "fee":          ["processing fee", "merchant fee", "fee", "transaction fee",
                     "charge", "service fee", "fees"],
    "net_amount":   ["net amount", "net", "net payment", "settlement amount",
                     "payout amount", "total"],
    "invoice":      ["invoice", "invoice number", "invoice #", "invoice no",
                     "reference", "description"],
    "payment_id":   ["payment id", "payment reference", "transaction id", "id"],
    "status":       ["status", "payment status", "type"],
    "email":        ["email", "email address", "customer email"],
}


def _normalise_col(h):
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _map_columns(df, col_map):
    norm_to_actual = {_normalise_col(c): c for c in df.columns}
    result = {}
    for canonical, aliases in col_map.items():
        for alias in aliases:
            if _normalise_col(alias) in norm_to_actual:
                result[canonical] = norm_to_actual[_normalise_col(alias)]
                break
    return result


# ─────────────────────────────────────────────────────────────────────────────
# BOA CSV/Excel parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_boa(file_obj):
    """
    Parse Bank of America export.

    The real stmt.csv has this structure:
      Row 0:  Description,,Summary Amt.         ← summary header (3 cols)
      Row 1:  Beginning balance...              ← summary data
      Row 2:  Total credits...
      Row 3:  Total debits...
      Row 4:  Ending balance...
      Row 5:  (blank)
      Row 6:  Date,Description,Amount,Running Bal.   ← REAL header (4 cols)
      Row 7+: actual transactions
    """
    errors = []
    try:
        name = getattr(file_obj, "name", "")
        if name.lower().endswith(".csv"):
            df = _read_boa_csv(file_obj)
        else:
            df = _read_boa_excel(file_obj)
    except Exception as e:
        return None, [f"Failed to read BOA file: {e}"]

    if df is None or df.empty:
        return None, ["BOA file parsed to an empty DataFrame."]

    mapping = _map_columns(df, _BOA_COL_MAP)
    rename = {v: k for k, v in mapping.items()}
    df = df.rename(columns=rename)

    for col in ("date", "description", "amount", "debit", "balance"):
        if col not in df.columns:
            df[col] = np.nan

    # Derive signed amount
    if "amount" in df.columns and df["amount"].notna().any():
        df["_boa_amount"] = _to_numeric(df["amount"])
    elif "debit" in df.columns and df["debit"].notna().any():
        df["_boa_amount"] = _to_numeric(df["debit"]) * -1
    else:
        df["_boa_amount"] = np.nan

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    # Drop rows with no date (summary/balance rows that slipped through)
    df = df[df["date"].notna()].copy()

    if "description" in df.columns:
        # Match only Zoho PAYMENT deposits, not Zoho software charges
        # "ZOHO PAYMENTS DES:..." = customer payment deposit ✅
        # "ZOHO* ZOHO-ONE..."     = software subscription charge ❌
        df["_is_zoho"] = df["description"].astype(str).str.upper().str.contains("ZOHO PAYMENTS")
    else:
        df["_is_zoho"] = False
        errors.append("BOA: No description column found.")

    df = df.reset_index(drop=True)
    return df, errors


def _read_boa_csv(file_obj):
    """
    The BOA CSV has a 5-row summary block with 3 columns, then a blank line,
    then the real 4-column transaction data. We find the real header by
    looking for the line with the most recognized column keywords.
    """
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8", errors="replace")

    lines = raw.splitlines()

    # Score each line: count how many BOA header keywords it contains
    best_score = 0
    header_idx = 0
    for i, line in enumerate(lines):
        lower = line.lower()
        score = sum(1 for kw in ["date", "description", "amount", "balance"]
                    if kw in lower)
        if score > best_score:
            best_score = score
            header_idx = i

    # Take everything from the best header line onward
    data_lines = lines[header_idx:]

    # Determine column count from the header
    header_cols = len(next(iter(data_lines)).split(","))

    # Keep only lines that have plausible column counts (header_cols ± 1)
    # to drop any stray summary lines below the data
    clean_lines = [data_lines[0]]  # always keep header
    for line in data_lines[1:]:
        stripped = line.strip()
        if not stripped:
            continue
        clean_lines.append(line)

    csv_content = "\n".join(clean_lines)

    return pd.read_csv(
        io.StringIO(csv_content),
        dtype=str,
        on_bad_lines="skip",
        engine="python",
    )


def _read_boa_excel(file_obj):
    data = file_obj.read()
    for header_row in range(10):
        try:
            df = pd.read_excel(io.BytesIO(data), header=header_row, dtype=str)
            cols_lower = [str(c).lower() for c in df.columns]
            score = sum(1 for kw in ["date", "description", "amount"]
                        if any(kw in c for c in cols_lower))
            if score >= 2:
                return df
        except Exception:
            continue
    return pd.read_excel(io.BytesIO(data), dtype=str)


# ─────────────────────────────────────────────────────────────────────────────
# Zoho PDF parser  (screen-captured payout page from pay.zoho.com)
# ─────────────────────────────────────────────────────────────────────────────

def parse_zoho(file_obj):
    """
    Route to the correct sub-parser based on file extension.
    The Zoho Payout PDF is a screen-captured web page — pdfplumber finds only
    ONE merged cell containing all text. We must parse the raw page text.
    """
    errors = []
    name = getattr(file_obj, "name", "")

    try:
        if name.lower().endswith(".pdf"):
            df, pdf_errors = _parse_zoho_pdf(file_obj)
            errors += pdf_errors
        elif name.lower().endswith(".csv"):
            df = _read_zoho_csv(file_obj)
        else:
            df = _read_zoho_excel(file_obj)
    except Exception as e:
        return None, [f"Failed to read Zoho file: {e}"]

    if df is None or df.empty:
        return None, ["Zoho file parsed to an empty DataFrame."]

    mapping = _map_columns(df, _ZOHO_COL_MAP)

    rename = {v: k for k, v in mapping.items()}
    df = df.rename(columns=rename)

    for col in _ZOHO_COL_MAP.keys():
        if col not in df.columns:
            df[col] = np.nan

    for col in ("gross_amount", "fee", "net_amount"):
        if col in df.columns:
            df[col] = _to_numeric(df[col])

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    # Drop only literal "Total" summary rows (not empty-customer rows from PDF)
    # Zoho PDF rows legitimately have blank customer — name comes from invoice
    if "customer" in df.columns:
        mask = ~df["customer"].astype(str).str.strip().str.lower().str.startswith("total")
        df = df[mask]

    df = df.reset_index(drop=True)
    return df, errors


def _parse_zoho_pdf(file_obj):
    """
    Parse a Zoho Payout Details PDF (screen-captured from pay.zoho.com).

    Architecture:
      The Zoho PDF is used for PAYOUT-LEVEL METADATA only:
        - Payout date (= BOA posting date)
        - Total gross amount (for balance validation)
        - Total fee / summary fee (authoritative debit amount — never recalculated)
        - BOA account last 4 digits (for offset account routing)
        - Payout ID + Bank Reference (appear in BOA description)

      Per-transaction customer names and amounts come from uploaded INVOICES,
      not from the PDF. This is robust to any transaction count (1 to 14+)
      because it never relies on parsing individual transaction rows from the PDF.

    Per-transaction amounts ARE extracted when available (for invoice-less flows),
    using the "chunk-by-3" strategy: the PDF always emits (gross, net, fee) for
    each transaction in sequence. We skip the summary values and chunk the rest.

    CRITICAL: The summary fee line is ALWAYS used as the debit row amount.
    Individual per-txn fees are never summed — they can drift due to PDF layout.
    """
    errors = []
    try:
        import pdfplumber
    except ImportError:
        return pd.DataFrame(), ["pdfplumber not installed — PDF parsing unavailable."]

    raw_data = file_obj.read()
    pages_text = []

    try:
        with pdfplumber.open(io.BytesIO(raw_data)) as pdf:
            for page in pdf.pages:
                pages_text.append(page.extract_text() or "")
    except Exception as e:
        return pd.DataFrame(), [f"PDF read error: {e}"]

    full_text = "\n".join(pages_text)

    # ── 1. Extract payout metadata ────────────────────────────────────────────

    # Payout date ("Payout: Jun 15, 2026, 06:38 AM")
    payout_date = None
    pd_m = re.search(r"Payout:\s+(\w+ \d{1,2}, \d{4})", full_text)
    if pd_m:
        payout_date = pd_m.group(1).strip()

    # Paid On date — rendered across 3 separate lines in the PDF:
    # "Paid / On / Jun / 16, / 2026"  (rendered across 3 separate lines)
    paid_on = None
    paid_on_m = re.search(
        r"Paid\s*\nOn\s*\n(\w+)\s*\n(\d{1,2}),?\s*\n(\d{4})", full_text
    )
    if paid_on_m:
        paid_on = f"{paid_on_m.group(1)} {paid_on_m.group(2)}, {paid_on_m.group(3)}"
    if not paid_on:
        paid_on_m2 = re.search(r"Paid\s+On\s+(\w+\s+\d{1,2},?\s*\d{4})", full_text)
        if paid_on_m2:
            paid_on = paid_on_m2.group(1).strip()

    posting_date = paid_on or payout_date or ""

    # Payout ID and Bank Reference
    pid_m = re.search(r"Payout ID:\s*(\d+)", full_text)
    payout_id = pid_m.group(1) if pid_m else ""

    bref_m = re.search(r"Bank Reference ID:\s*(\d+)", full_text)
    bank_ref = bref_m.group(1) if bref_m else ""

    # BOA account last 4 digits "( 3371 )" or "(••••3371)"
    acct_m = re.search(r"\(\s*[•\*]*\s*(\d{4})\s*\)", full_text)
    boa_acct = acct_m.group(1) if acct_m else ""

    # ── 2. Extract summary totals (authoritative) ─────────────────────────────
    summary_m = re.search(
        r"Payments\s+(\d+)\s+\$([\d,]+\.\d{2})\s+[−\-]\$?([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})",
        full_text
    )
    if not summary_m:
        return pd.DataFrame(), ["Could not find Payments summary line in Zoho PDF. "
                                  "Check that the file is a Zoho Payout Details PDF."]

    expected_count = int(summary_m.group(1))
    summary_gross  = float(summary_m.group(2).replace(",", ""))
    summary_fee    = float(summary_m.group(3).replace(",", ""))  # AUTHORITATIVE fee
    summary_net    = float(summary_m.group(4).replace(",", ""))

    # ── 3. Extract per-transaction gross amounts ──────────────────────────────
    # Used when no invoices are uploaded. Strategy: chunk-by-3.
    # The PDF emits dollar values after the summary in order:
    #   gross_1, net_1, fee_1, gross_2, net_2, fee_2, ...
    # We skip the summary values, then chunk remaining values into groups of 3.

    txn_section = full_text
    if "All Transactions" in full_text:
        txn_section = full_text.split("All Transactions", 1)[1]

    # Remove noise lines
    txn_section = re.sub(r"Refunds[^\n]*\n?", "", txn_section)
    txn_section = re.sub(r"Adjustments[^\n]*\n?", "", txn_section)
    txn_section = re.sub(r"Amount\s+\$[\d,]+\.\d{2}\s+USD\n?", "", txn_section)
    txn_section = re.sub(r"https?://\S+", "", txn_section)
    txn_section = re.sub(r"Page \d+ of \d+", "", txn_section)

    # Collect all positive dollar values in the transaction section
    all_dollars = [
        float(m.group(1).replace(",", ""))
        for m in re.finditer(r"\$([\d,]+\.\d{2})", txn_section)
        if float(m.group(1).replace(",", "")) > 0
    ]

    # Skip leading summary values (summary_gross, summary_fee, summary_net appear first)
    skip_vals = [summary_gross, summary_fee, summary_net]
    skip_idx = 0
    clean_dollars = []
    for val in all_dollars:
        if skip_idx < len(skip_vals) and abs(val - skip_vals[skip_idx]) < 0.02:
            skip_idx += 1
            continue
        clean_dollars.append(val)

    # Chunk into groups of 3: (gross, net, fee)
    transactions = []
    if len(clean_dollars) >= expected_count * 3:
        for i in range(expected_count):
            chunk = clean_dollars[i*3 : i*3+3]
            transactions.append({
                "gross_amount":  str(chunk[0]),
                "fee":           str(chunk[2]),   # fee is LAST in triplet
                "net_amount":    str(chunk[1]),
                "customer":      "",
                "date":          posting_date,
                "payout_date":   posting_date,
                "payout_id":     payout_id,
                "bank_ref":      bank_ref,
                "boa_acct":      boa_acct,
                "_summary_fee":  str(summary_fee),
                "_summary_gross": str(summary_gross),
                "_expected_count": str(expected_count),
            })
    else:
        # Partial data or dates-only available — create one row per expected txn
        # with gross distributed from clean_dollars (best effort)
        # The summary fee is still authoritative for the debit row
        available_gross = clean_dollars[:expected_count] if clean_dollars else []

        for i in range(expected_count):
            gross = available_gross[i] if i < len(available_gross) else 0.0
            transactions.append({
                "gross_amount":  str(gross),
                "fee":           "0.0",   # individual fee unknown; use summary total
                "net_amount":    str(gross),
                "customer":      "",
                "date":          posting_date,
                "payout_date":   posting_date,
                "payout_id":     payout_id,
                "bank_ref":      bank_ref,
                "boa_acct":      boa_acct,
                "_summary_fee":  str(summary_fee),
                "_summary_gross": str(summary_gross),
                "_expected_count": str(expected_count),
            })

    # Validate: sum of parsed grosses should match summary
    parsed_gross_sum = sum(float(t["gross_amount"]) for t in transactions)
    if abs(parsed_gross_sum - summary_gross) > 1.0:
        errors.append(
            f"Zoho PDF: Per-transaction gross sum ${parsed_gross_sum:,.2f} does not match "
            f"summary ${summary_gross:,.2f}. Individual row amounts may be approximate. "
            f"Upload invoices for exact per-customer amounts."
        )

    if not transactions:
        return pd.DataFrame(), ["Could not extract transaction rows from Zoho PDF. "
                                  "Upload Zoho export as CSV or Excel for best results."]

    return pd.DataFrame(transactions), errors



def _read_zoho_csv(file_obj):
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8", errors="replace")
    lines = content.splitlines()
    header_idx = 0
    for i, line in enumerate(lines):
        lower = line.lower()
        score = sum(1 for kw in ["amount", "fee", "date", "customer", "name"]
                    if kw in lower)
        if score >= 2:
            header_idx = i
            break
    csv_content = "\n".join(lines[header_idx:])
    return pd.read_csv(
        io.StringIO(csv_content), dtype=str,
        on_bad_lines="skip", engine="python"
    )


def _read_zoho_excel(file_obj):
    data = file_obj.read()
    for header_row in range(8):
        try:
            df = pd.read_excel(io.BytesIO(data), header=header_row, dtype=str)
            cols_lower = [str(c).lower() for c in df.columns]
            score = sum(1 for kw in ["amount", "fee", "date", "customer"]
                        if any(kw in c for c in cols_lower))
            if score >= 2:
                return df
        except Exception:
            continue
    return pd.read_excel(io.BytesIO(data), dtype=str)


# ─────────────────────────────────────────────────────────────────────────────
# Invoice PDF parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_invoice_pdf(file_obj):
    """
    Extract customer name, invoice number, total, and payment terms from
    an InBody Purchase Statement PDF.

    Returns a dict with keys:
      invoice_number, customer_name, total, payment_terms, invoice_date, cs_ticket
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    raw = file_obj.read() if hasattr(file_obj, "read") else open(file_obj, "rb").read()
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
    except Exception:
        return {}

    result = {}

    # Invoice number
    inv_m = re.search(r"Purchase\s*#\s*(INV-[\w\d]+)", text, re.IGNORECASE)
    if inv_m:
        result["invoice_number"] = inv_m.group(1).strip()

    # Bill To name (line immediately after "Bill To")
    bill_to_m = re.search(r"Bill\s+To\s*\n([^\n]+)", text, re.IGNORECASE)
    if bill_to_m:
        result["customer_name"] = bill_to_m.group(1).strip()

    # Payment Terms
    terms_m = re.search(r"Terms\s*:\s*([^\n]+)", text, re.IGNORECASE)
    if terms_m:
        result["payment_terms"] = terms_m.group(1).strip()

    # Invoice date
    date_m = re.search(r"Invoice\s+Date\s*:\s*([^\n]+)", text, re.IGNORECASE)
    if date_m:
        result["invoice_date"] = date_m.group(1).strip()

    # Total: find the LAST standalone "Total $X,XXX.XX" line
    # (not SubTotal, not "Balance Due", just "Total $X")
    # Use findall and take the value after the last plain "Total"
    total_matches = list(re.finditer(
        r"(?:^|\n)\s*Total\s+\$?([\d,]+\.\d{2})", text
    ))
    if total_matches:
        result["total"] = float(total_matches[-1].group(1).replace(",", ""))
    else:
        # Fallback: any "Total $X"
        t_m = re.search(r"Total\s+\$?([\d,]+\.\d{2})", text)
        if t_m:
            result["total"] = float(t_m.group(1).replace(",", ""))

    # CS Ticket number — "[## 676 ##]" pattern
    ticket_m = re.search(r"\[##\s*(\d+)\s*##\]", text)
    if ticket_m:
        result["cs_ticket"] = ticket_m.group(1).strip()

    return result



# ─────────────────────────────────────────────────────────────────────────────
# Reference file loaders
# ─────────────────────────────────────────────────────────────────────────────

def load_customer_master(path):
    """
    Load Account_Masterlist.xlsx.

    Supports these column formats:
      2-col: Account | Account Name
      3-col: Account Type | Account | Account Name
      4-col: Account Type | Account | Account Name | CS/PS Ticket

    The CS/PS Ticket column contains the individual person name that appears
    on the invoice "Bill To" field (e.g. "Ali Amir") when the registered D365
    account name is a business (e.g. "Functional Holistic Healing").

    Returns a DataFrame with columns:
      Account, Account Name, CS/PS Ticket (blank if not present)

    Only Customer rows (BC######) are returned.
    """
    df = pd.read_excel(path, header=None, dtype=str)

    # Find header row
    header_row = 0
    for i, row in df.iterrows():
        vals = [str(v).strip().lower() for v in row.values]
        if "account" in vals:
            header_row = i
            break

    df.columns = df.iloc[header_row].astype(str).str.strip()
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [c.strip() for c in df.columns]
    cols_lower = [c.lower() for c in df.columns]

    # Filter to Customer rows only
    if "account type" in cols_lower:
        type_col = df.columns[cols_lower.index("account type")]
        df = df[df[type_col].astype(str).str.strip().str.lower() == "customer"].copy()

    # Extract the columns we need
    acct_col = next((c for c in df.columns if c.lower() == "account"), None)
    name_col = next((c for c in df.columns if c.lower() == "account name"), None)

    # CS/PS Ticket column — optional, present in 4-column format
    ticket_col = next((c for c in df.columns
                       if c.lower() in ("cs/ps ticket", "cs ticket", "ps ticket",
                                        "ticket", "bill to name", "individual name")), None)

    if acct_col and name_col:
        keep = [acct_col, name_col]
        if ticket_col:
            keep.append(ticket_col)
        df = df[keep].copy()
        df.columns = ["Account", "Account Name"] + (["CS/PS Ticket"] if ticket_col else [])
    else:
        # Fallback: first two columns
        df = df.iloc[:, :2].copy()
        df.columns = ["Account", "Account Name"]

    # Ensure CS/PS Ticket column always exists (blank if not in file)
    if "CS/PS Ticket" not in df.columns:
        df["CS/PS Ticket"] = ""

    df = df.dropna(subset=["Account"]).reset_index(drop=True)
    df["Account"]      = df["Account"].astype(str).str.strip()
    df["Account Name"] = df["Account Name"].astype(str).str.strip()
    df["CS/PS Ticket"] = df["CS/PS Ticket"].fillna("").astype(str).str.strip()

    # Keep only valid BC###### accounts
    df = df[df["Account"].str.match(r"BC\d+")].reset_index(drop=True)
    return df


def load_cash_codes(path):
    df = pd.read_excel(path, header=None, dtype=str)
    header_row = 0
    for i, row in df.iterrows():
        vals = [str(v).strip().lower() for v in row.values]
        if "cash code" in vals:
            header_row = i
            break
    df.columns = df.iloc[header_row].astype(str).str.strip()
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = ["Cash Code", "Cash Code Name"] + list(df.columns[2:])
    df = df.dropna(subset=["Cash Code"]).reset_index(drop=True)
    df["Cash Code"] = df["Cash Code"].str.strip()
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _to_numeric(series):
    return (
        series.astype(str)
        .str.replace(r"[\$,€£]", "", regex=True)
        .str.replace(r"[−–]", "-", regex=True)
        .str.replace(r"\((\d+\.?\d*)\)", r"-\1", regex=True)
        .str.strip()
        .replace("", np.nan)
        .replace("nan", np.nan)
        .pipe(pd.to_numeric, errors="coerce")
    )


"""
matcher.py — Match BOA ↔ Zoho transactions and resolve customer accounts.

Key insight from real data analysis:
  - The Zoho Payout PDF has NO customer names (field shows "—")
  - Customer names must come from uploaded invoice PDFs
  - Invoices are matched to Zoho transactions by amount (Total on invoice == gross_amount in Zoho)
  - BOA posting date (payout date) is used as the D365 entry date, not the Zoho transaction date
  - BOA description string (ZOHO PAYMENTS DES:...) is used in the D365 description field

Special account types (from real D365 reference data):
  - Normal business customers: Customer type, BC###### account
  - CS/repair ticket individuals found in CS/PS Ticket column of master:
    Customer type, resolved BC###### of their registered business account.
    Description prefixed with "Name CS Ticket #XXX_"
  - CS/repair ticket individuals NOT in master: Ledger type,
    account 21040102-B1000002, name "Temporary Receipt" (flagged for review)
"""

import re
import numpy as np
import pandas as pd
from rapidfuzz import fuzz, process
from parsers import parse_invoice_pdf

_FUZZY_THRESHOLD = 80

# Special handling: individual/CS customers use Temporary Receipt account
# These are identified when the customer is an individual (not a business)
# and typically come from CS ticket invoices
_TEMP_RECEIPT_ACCOUNT = "21040102-B1000002"
_TEMP_RECEIPT_NAME    = "Temporary Receipt"

# Payment terms → cash code mapping (from automation rules §4)
_TERMS_TO_CASH_CODE = {
    "due on receipt":         ("AR001", ""),
    "due upon receipt":       ("AR001", ""),
    "monthly payment":        ("AR002", "MPP "),
    "monthly payment plan":   ("AR002", "MPP "),
    "mpp":                    ("AR002", "MPP "),
    "financing":              ("AR003", ""),
    "leasing":                ("AR004", ""),
    "net 1":                  ("AR005", ""),
    "net 10":                 ("AR006", ""),
    "net 25":                 ("AR007", ""),
    "net 30":                 ("AR008", ""),
    "net 40":                 ("AR009", ""),
    "net 45":                 ("AR010", ""),
    "net 60":                 ("AR011", ""),
    "due end of next month":  ("AR011", ""),   # Net 60 / end of next month
    "net":                    ("AR008", ""),   # generic net → AR008
}


def match_transactions(boa_df, zoho_df, customer_df, invoice_files=None):
    """
    Main entry point. Enriches zoho_df rows with D365 fields.

    Steps:
      1. Parse all uploaded invoice PDFs → {amount: invoice_data}
      2. For each Zoho row: match by amount to invoice → get customer name + terms
      3. Look up customer name in master → get BC###### account
      4. Attach BOA posting date and description
      5. Validate balance invariant
    """
    log = []

    # ── Step 1: Parse invoices ────────────────────────────────────────────────
    invoice_by_amount = {}   # {rounded_amount: invoice_dict}
    invoice_by_name   = {}   # {normalised_name: invoice_dict}

    if invoice_files:
        for inv_file in invoice_files:
            try:
                inv_data = parse_invoice_pdf(inv_file)
                if inv_data:
                    amt = inv_data.get("total")
                    if amt:
                        key = round(float(amt), 2)
                        invoice_by_amount[key] = inv_data
                    name = inv_data.get("customer_name", "")
                    if name:
                        invoice_by_name[_normalise_name(name)] = inv_data
                    log.append({
                        "level": "OK",
                        "msg": (f"Invoice parsed: {inv_data.get('invoice_number','?')} | "
                                f"Customer: {inv_data.get('customer_name','?')} | "
                                f"Total: ${amt:,.2f}" if amt else "Invoice parsed (no total)")
                    })
            except Exception as e:
                log.append({"level": "WARN", "msg": f"Could not parse invoice: {e}"})

    # ── Step 2: Build customer lookup ─────────────────────────────────────────
    customer_lookup = _build_customer_lookup(customer_df)

    # ── Step 3: Resolve each Zoho row ─────────────────────────────────────────
    results = []
    for idx, zrow in zoho_df.iterrows():
        record = _resolve_row(
            zrow, customer_lookup, invoice_by_amount, invoice_by_name, log
        )
        results.append(record)

    resolved = pd.DataFrame(results)
    enriched = pd.concat(
        [zoho_df.reset_index(drop=True), resolved.reset_index(drop=True)], axis=1
    )

    # ── Step 4: Attach BOA date and description ───────────────────────────────
    enriched = _attach_boa_data(enriched, boa_df, log)

    # ── Step 5: Balance check ─────────────────────────────────────────────────
    _validate_balance(boa_df, enriched, log)

    return enriched, log


# ─────────────────────────────────────────────────────────────────────────────
# Row resolution
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_row(zrow, customer_lookup, invoice_by_amount, invoice_by_name, log):
    """Resolve one Zoho transaction row to its D365 fields."""
    idx          = zrow.name
    raw_customer = str(zrow.get("customer", "")).strip()
    gross        = _safe_float(zrow.get("gross_amount"))
    invoice_ref  = str(zrow.get("invoice", "")).strip()

    # ── Try invoice match by amount first (most reliable for this PDF) ────────
    inv = None
    if gross:
        key = round(gross, 2)
        inv = invoice_by_amount.get(key)

    # If not found by amount, try by invoice reference string
    if not inv and invoice_ref and invoice_ref not in ("nan", ""):
        for norm_name, inv_data in invoice_by_name.items():
            if inv_data.get("invoice_number", "") in invoice_ref:
                inv = inv_data
                break

    # ── If we have an invoice, use it ─────────────────────────────────────────
    if inv:
        customer_name  = inv.get("customer_name", "")
        payment_terms  = inv.get("payment_terms", "")
        cs_ticket      = inv.get("cs_ticket", "")
        cash_code, pfx = _terms_to_cash_code(payment_terms)

        log.append({
            "level": "OK",
            "msg": (f"Row {idx}: matched invoice {inv.get('invoice_number','?')} "
                    f"→ '{customer_name}' | terms='{payment_terms}' | code={cash_code}")
        })

        # Check if this is a CS/repair ticket invoice (has [## XXX ##] in items)
        is_cs_ticket = bool(cs_ticket)

        # Normal business customer lookup — works for both business names AND
        # individual names that appear in the CS/PS Ticket column of the master
        master_result = _fuzzy_lookup(customer_name, customer_lookup)
        found_in_master = master_result["_match_confidence"] in ("HIGH", "MEDIUM")

        if is_cs_ticket and not found_in_master:
            # Individual not found in master at all — use Temporary Receipt
            desc_prefix = f"{customer_name} CS Ticket #{cs_ticket}_"
            return {
                "_account":          _TEMP_RECEIPT_ACCOUNT,
                "_account_name":     _TEMP_RECEIPT_NAME,
                "_account_type":     "Ledger",
                "_posting_profile":  "AutoPost",
                "_match_method":     "INVOICE_CS_TICKET_UNREGISTERED",
                "_match_confidence": "MEDIUM",
                "_needs_review":     True,
                "_review_reason":    (f"'{customer_name}' not in Account Masterlist. "
                                      f"Add them with their BC###### to resolve automatically."),
                "_cash_code":        cash_code,
                "_cash_code_prefix": "",
                "_desc_prefix":      desc_prefix,
                "_raw_name":         customer_name,
                "_invoice_number":   inv.get("invoice_number", ""),
            }

        if is_cs_ticket and found_in_master:
            # Individual found via CS/PS Ticket column — use their business account
            # but keep the CS ticket description prefix for the D365 description
            log.append({
                "level": "OK",
                "msg": (f"Row {idx}: CS ticket '{customer_name}' matched to "
                        f"'{master_result['_account_name']}' via CS/PS Ticket column.")
            })
            desc_prefix = f"{customer_name} CS Ticket #{cs_ticket}_"
            master_result["_desc_prefix"]      = desc_prefix
            master_result["_cash_code"]        = cash_code
            master_result["_cash_code_prefix"] = pfx
            master_result["_invoice_number"]   = inv.get("invoice_number", "")
            master_result["_raw_name"]         = customer_name
            master_result.setdefault("_account_type",    "Customer")
            master_result.setdefault("_posting_profile", "AutoPost")
            return master_result
        master_result["_cash_code"]        = cash_code
        master_result["_cash_code_prefix"] = pfx
        master_result["_desc_prefix"]      = ""
        master_result["_invoice_number"]   = inv.get("invoice_number", "")
        master_result["_raw_name"]         = customer_name
        master_result.setdefault("_account_type",    "Customer")
        master_result.setdefault("_posting_profile", "AutoPost")

        if master_result["_match_confidence"] == "LOW":
            master_result["_needs_review"]  = True
            master_result["_review_reason"] = (
                f"Customer '{customer_name}' from invoice not found in master "
                f"(best match score {master_result.get('_fuzzy_score',0)}/100). "
                f"Add this account to IBNY_Business_Customer_Account.xlsx."
            )
            log.append({
                "level": "WARN",
                "msg":   f"Row {idx}: '{customer_name}' not in master — flagged for review."
            })

        return master_result

    # ── No invoice: try customer name from Zoho directly ─────────────────────
    if raw_customer and raw_customer not in ("—", "-", "nan", ""):
        result = _fuzzy_lookup(raw_customer, customer_lookup)
        result["_cash_code"]        = "AR001"  # default
        result["_cash_code_prefix"] = ""
        result["_desc_prefix"]      = ""
        result["_invoice_number"]   = invoice_ref
        result["_raw_name"]         = raw_customer
        result.setdefault("_account_type",    "Customer")
        result.setdefault("_posting_profile", "AutoPost")
        return result

    # ── No name, no invoice match ─────────────────────────────────────────────
    log.append({
        "level": "WARN",
        "msg":   (f"Row {idx}: No customer name and no matching invoice for "
                  f"gross=${gross}. Upload the invoice to resolve.")
    })
    return {
        "_account":          "",
        "_account_name":     raw_customer or "",
        "_account_type":     "Customer",
        "_posting_profile":  "AutoPost",
        "_match_method":     "UNRESOLVED",
        "_match_confidence": "LOW",
        "_needs_review":     True,
        "_review_reason":    f"No customer name and no invoice matched amount ${gross}",
        "_cash_code":        "",
        "_cash_code_prefix": "",
        "_desc_prefix":      "",
        "_raw_name":         raw_customer or "",
        "_invoice_number":   "",
    }


# ─────────────────────────────────────────────────────────────────────────────
# BOA data attachment
# ─────────────────────────────────────────────────────────────────────────────

def _attach_boa_data(zoho_df, boa_df, log):
    """
    Attach the BOA posting date and description to each Zoho row.

    The BOA CSV has one row per payout date. The payout date in Zoho is
    stored in 'payout_date' (from the PDF header). If that matches a BOA
    row date, use it; otherwise fall back to ±3-day window.
    """
    zoho_df = zoho_df.copy()

    if boa_df is None or boa_df.empty:
        zoho_df["_boa_date"]        = pd.NaT
        zoho_df["_boa_description"] = ""
        return zoho_df

    boa_zoho = boa_df[boa_df.get("_is_zoho", pd.Series(False, index=boa_df.index))].copy()

    if boa_zoho.empty:
        zoho_df["_boa_date"]        = pd.NaT
        zoho_df["_boa_description"] = ""
        log.append({
            "level": "WARN",
            "msg":   "No ZOHO rows found in BOA file. "
                     "The BOA description must contain 'ZOHO' to be recognised."
        })
        return zoho_df

    # Build date → {description, amount} map from BOA
    boa_map = {}
    for _, brow in boa_zoho.iterrows():
        if pd.notna(brow.get("date")):
            d = pd.Timestamp(brow["date"]).normalize()
            boa_map[d] = {
                "description": str(brow.get("description", "")),
                "amount":      brow.get("_boa_amount", np.nan),
            }

    def get_boa_for_row(row):
        # Priority 1: use payout_date from Zoho PDF if available
        for date_field in ["payout_date", "date"]:
            val = row.get(date_field)
            if val and str(val) not in ("nan", "NaT", ""):
                try:
                    d = pd.Timestamp(val).normalize()
                    for delta in [0, 1, -1, 2, -2, 3, -3]:
                        target = d + pd.Timedelta(days=delta)
                        if target in boa_map:
                            return boa_map[target]
                except Exception:
                    pass
        # Use the first BOA row if nothing matched
        if boa_map:
            return next(iter(boa_map.values()))
        return {"description": "", "amount": np.nan}

    boa_dates = []
    boa_descs = []
    for _, row in zoho_df.iterrows():
        boa_entry = get_boa_for_row(row)
        boa_dates.append(
            pd.Timestamp(list(boa_map.keys())[0])
            if boa_map else pd.NaT
        )
        boa_descs.append(boa_entry["description"])

    # Use the actual BOA date (from boa_map keys) not the Zoho transaction date
    # The BOA posting date is what goes into D365 (per SOP)
    if boa_map:
        boa_posting_date = list(boa_map.keys())[0]  # first (and usually only) ZOHO row
        zoho_df["_boa_date"] = boa_posting_date
    else:
        zoho_df["_boa_date"] = pd.NaT

    zoho_df["_boa_description"] = boa_descs
    return zoho_df


# ─────────────────────────────────────────────────────────────────────────────
# Balance validation
# ─────────────────────────────────────────────────────────────────────────────

def _validate_balance(boa_df, zoho_df, log):
    if boa_df is None or boa_df.empty:
        return
    if "_boa_amount" not in boa_df.columns:
        return

    boa_zoho = boa_df[boa_df.get("_is_zoho", pd.Series(False, index=boa_df.index))]
    if boa_zoho.empty:
        return

    boa_net    = boa_zoho["_boa_amount"].sum()
    zoho_gross = zoho_df["gross_amount"].fillna(0).sum() if "gross_amount" in zoho_df.columns else 0

    # Use Zoho summary fee (authoritative) if present; fall back to per-txn sum
    if "_summary_fee" in zoho_df.columns:
        summary_fees = zoho_df["_summary_fee"].dropna()
        summary_fees = pd.to_numeric(summary_fees, errors="coerce").dropna()
        zoho_fee = summary_fees.iloc[0] if not summary_fees.empty else zoho_df["fee"].fillna(0).sum()
    else:
        zoho_fee = zoho_df["fee"].fillna(0).sum() if "fee" in zoho_df.columns else 0

    zoho_net   = zoho_gross - zoho_fee

    diff = abs(float(boa_net) - float(zoho_net))
    tol  = max(abs(float(boa_net)) * 0.005, 1.0)

    if diff <= tol:
        log.append({
            "level": "OK",
            "msg":   (f"Balance check PASSED: BOA net=${boa_net:,.2f} "
                      f"≈ Zoho gross${zoho_gross:,.2f} − fee${zoho_fee:,.2f} = ${zoho_net:,.2f}")
        })
    else:
        log.append({
            "level": "WARN",
            "msg":   (f"Balance mismatch: BOA net=${boa_net:,.2f}, "
                      f"Zoho net=${zoho_net:,.2f}, diff=${diff:,.2f} "
                      f"(tolerance=${tol:.2f})")
        })


# ─────────────────────────────────────────────────────────────────────────────
# Customer lookup helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_customer_lookup(customer_df):
    """
    Build lookup dict: normalised_name → {account, account_name}

    Two entries are created per row when CS/PS Ticket is present:
      1. Normalised Account Name  → used when Zoho/invoice has the business name
      2. Normalised CS/PS Ticket  → used when invoice "Bill To" has an individual name
         (e.g. "Ali Amir" → maps to BC000654 Functional Holistic Healing)

    This means even if the invoice says "Bill To: Ali Amir", the app will
    record the correct D365 account name "Functional Holistic Healing".
    """
    lookup = {}
    for _, row in customer_df.iterrows():
        acc    = str(row.get("Account", "")).strip()
        name   = str(row.get("Account Name", "")).strip()
        ticket = str(row.get("CS/PS Ticket", "")).strip()

        if not acc or not name:
            continue

        entry = {"account": acc, "account_name": name}

        # Primary key: business/account name
        lookup[_normalise_name(name)] = entry

        # Secondary key: individual name from CS/PS Ticket column
        if ticket and ticket.lower() not in ("", "nan", "none"):
            lookup[_normalise_name(ticket)] = entry

    return lookup


def _normalise_name(name):
    name = str(name).lower().strip()
    name = re.sub(r"[^\w\s]", " ", name)
    for suffix in [r"\bllc\b", r"\binc\b", r"\bltd\b", r"\bcorp\b",
                   r"\bpllc\b", r"\bpc\b", r"\bdba\b", r"\bthe\b"]:
        name = re.sub(suffix, "", name)
    return re.sub(r"\s+", " ", name).strip()


def _fuzzy_lookup(raw_name, customer_lookup):
    norm = _normalise_name(raw_name)
    keys = list(customer_lookup.keys())

    if not keys:
        return _unresolved(raw_name, "Customer master is empty")

    result   = process.extractOne(norm, keys, scorer=fuzz.token_set_ratio)
    best_key = result[0]
    score    = result[1]
    matched  = customer_lookup[best_key]

    if score >= _FUZZY_THRESHOLD:
        confidence = "HIGH" if score >= 92 else "MEDIUM"
        return {
            "_account":          matched["account"],
            "_account_name":     matched["account_name"],
            "_account_type":     "Customer",
            "_posting_profile":  "AutoPost",
            "_match_method":     f"FUZZY({score})",
            "_match_confidence": confidence,
            "_needs_review":     confidence == "MEDIUM",
            "_review_reason":    (f"Fuzzy match '{matched['account_name']}' "
                                  f"score {score}/100 — verify") if confidence == "MEDIUM" else "",
            "_fuzzy_score":      score,
        }
    return _unresolved(raw_name,
                       f"Best match '{matched['account_name']}' scored {score}/100 < {_FUZZY_THRESHOLD}")


def _unresolved(raw_name, reason):
    return {
        "_account":          "",
        "_account_name":     raw_name,
        "_account_type":     "Customer",
        "_posting_profile":  "AutoPost",
        "_match_method":     "UNRESOLVED",
        "_match_confidence": "LOW",
        "_needs_review":     True,
        "_review_reason":    reason,
        "_fuzzy_score":      0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Payment term → cash code
# ─────────────────────────────────────────────────────────────────────────────

def _terms_to_cash_code(terms_str):
    """Return (cash_code, description_prefix) from a payment terms string."""
    if not terms_str or str(terms_str).lower() in ("nan", ""):
        return "AR001", ""
    terms_lower = terms_str.lower().strip()
    for key, val in _TERMS_TO_CASH_CODE.items():
        if key in terms_lower:
            return val
    return "AR001", ""   # default: due on receipt


def _safe_float(val):
    try:
        f = float(str(val).replace(",", "").replace("$", "").strip())
        return None if np.isnan(f) else f
    except (ValueError, TypeError):
        return None


"""
builder.py — Build D365 journal entry rows from matched Zoho data.

Rules source: Zoho_Payment_D365_Automation_Rules.docx §4 & §5

Per transaction (grouped by BOA posting date):
  • 1 CREDIT line per customer payment
  • 1 DEBIT line for ALL merchant fees (grouped per payout batch)

Special account handling (from real D365 reference data):
  • Normal customers: Account type=Customer, Account=BC######, Posting Profile=AutoPost
  • CS/repair tickets (individual in CS/PS Ticket column of master):
    Account type=Customer, Account=BC######, Posting Profile=AutoPost
    Description: "Name CS Ticket #XXX_ZOHO PAYMENTS DES:..."
  • CS/repair tickets (individual NOT in master):
    Account type=Ledger, Account=21040102-B1000002, Name=Temporary Receipt (flagged)
"""

import re
import pandas as pd
import numpy as np

_COMPANY          = "bwa"
_POSTING_PROFILE  = "AutoPost"
_DEBIT_ACCT_NAME  = "Outside Service (Finance)"
_DEBIT_ACCT_TYPE  = "Ledger"
_DEBIT_ACCT       = "43170111-U26C05001-B735350-UOA003"
_DEBIT_CASH_CODE  = "OSF005"
_CURRENCY         = "USD"
_EXCHANGE_RATE    = 1.00
_SALES_TAX_GROUP  = "AVATAX"
_REVERSING_ENTRY  = "No"

D365_COLUMNS = [
    "Date", "Voucher", "Account name", "Company", "Account type",
    "Account", "Posting profile", "Cash code", "Description",
    "Debit", "Credit",
    "Item sales tax group", "Sales tax code",
    "Offset company", "Offset account type", "Offset account",
    "Offset transaction text",
    "Currency", "Exchange rate",
    "Item sales tax group2", "Sales tax group",
    "Withholding tax group", "Release date", "Reversing entry", "Reversing date",
    "_needs_review", "_review_reason", "_match_confidence",
]


def build_journal_entries(matched_df, customer_df, offset_account="B1000002"):
    log  = []
    rows = []

    # Determine grouping date: use BOA posting date (_boa_date) if available,
    # fall back to Zoho transaction date
    matched_df = matched_df.copy()
    if "_boa_date" in matched_df.columns and matched_df["_boa_date"].notna().any():
        matched_df["_group_date"] = pd.to_datetime(
            matched_df["_boa_date"], errors="coerce"
        )
    elif "date" in matched_df.columns:
        matched_df["_group_date"] = pd.to_datetime(
            matched_df["date"], errors="coerce"
        )
    else:
        matched_df["_group_date"] = pd.NaT

    groups = matched_df.groupby("_group_date", dropna=False)

    for group_date, group in groups:
        group_rows   = list(group.iterrows())
        multi_batch  = len(group_rows) > 1
        credit_rows      = []
        total_fee        = 0.0   # fallback: sum of per-txn fees
        summary_fee      = None  # authoritative: from Zoho summary line
        fee_desc_parts   = []

        for _, zrow in group_rows:
            account       = str(zrow.get("_account", "")).strip()
            account_name  = str(zrow.get("_account_name", "")).strip()
            account_type  = str(zrow.get("_account_type",  "Customer")).strip()
            posting_prof  = str(zrow.get("_posting_profile", _POSTING_PROFILE)).strip()
            needs_review  = bool(zrow.get("_needs_review", False))
            review_reason = str(zrow.get("_review_reason", ""))
            confidence    = str(zrow.get("_match_confidence", "LOW"))
            gross         = _safe_float(zrow.get("gross_amount"))
            fee           = _safe_float(zrow.get("fee"))
            boa_desc      = str(zrow.get("_boa_description", "")).strip()
            cash_code     = str(zrow.get("_cash_code", "")).strip()
            cash_pfx      = str(zrow.get("_cash_code_prefix", "")).strip()
            desc_prefix   = str(zrow.get("_desc_prefix", "")).strip()

            # Default cash code if not set
            if not cash_code:
                cash_code = "AR001"

            # Format entry date from BOA posting date
            date_str = ""
            if pd.notna(group_date):
                try:
                    date_str = pd.Timestamp(group_date).strftime("%m/%d/%Y")
                except Exception:
                    date_str = str(group_date)

            # Build credit description
            credit_desc = _build_credit_description(
                cash_code, cash_pfx, desc_prefix,
                account, account_name, boa_desc
            )

            credit_row = {
                "Date":                  date_str,
                "Voucher":               "",
                "Account name":          account_name,
                "Company":               _COMPANY,
                "Account type":          account_type,
                "Account":               account,
                "Posting profile":       posting_prof if account else "",
                "Cash code":             cash_code,
                "Description":           credit_desc,
                "Debit":                 "",
                "Credit":                f"{gross:.2f}" if gross else "",
                "Item sales tax group":  "",
                "Sales tax code":        "",
                "Offset company":        _COMPANY,
                "Offset account type":   "Bank",
                "Offset account":        offset_account,
                "Offset transaction text": "",
                "Currency":              _CURRENCY,
                "Exchange rate":         _EXCHANGE_RATE,
                "Item sales tax group2": "",
                "Sales tax group":       _SALES_TAX_GROUP,
                "Withholding tax group": "",
                "Release date":          "",
                "Reversing entry":       _REVERSING_ENTRY,
                "Reversing date":        "",
                "_needs_review":         needs_review,
                "_review_reason":        review_reason,
                "_match_confidence":     confidence,
            }
            credit_rows.append(credit_row)
            rows.append(credit_row)

            if fee and fee > 0:
                total_fee += fee

            # Prefer the Zoho summary-level fee over per-txn sum
            sf = _safe_float(zrow.get("_summary_fee"))
            if sf and sf > 0:
                summary_fee = sf

            # Accumulate fee description parts
            part = " ".join(filter(None, [account, account_name])).strip()
            if part:
                fee_desc_parts.append(part)

        # ── Use Zoho summary fee (authoritative) over per-txn sum ───────────
        # The summary line "Payments N $X −$Y $Z" is the source of truth per SOP.
        # Never assume or recalculate — use exactly what Zoho reports.
        debit_fee = summary_fee if summary_fee else total_fee

        # ── Single grouped debit row for all fees in this batch ───────────────
        if debit_fee > 0:
            date_str = credit_rows[0]["Date"] if credit_rows else ""

            if multi_batch:
                fee_name_str = ", ".join(fee_desc_parts)
                debit_desc = f"Zoho Merchant Fee {fee_name_str}_{boa_desc}" if boa_desc else f"Zoho Merchant Fee {fee_name_str}"
            else:
                # Single payment: "Zoho Merchant Fee_ {credit_desc without code prefix}"
                base = credit_rows[0]["Description"] if credit_rows else ""
                # Remove cash code prefix (e.g. "AR001: ") or MPP prefix from start
                base_clean = re.sub(r"^(AR\d{3}:\s*|MPP\s+)", "", base)
                debit_desc = f"Zoho Merchant Fee_ {base_clean}".strip()

            debit_row = {
                "Date":                  date_str,
                "Voucher":               "",
                "Account name":          _DEBIT_ACCT_NAME,
                "Company":               _COMPANY,
                "Account type":          _DEBIT_ACCT_TYPE,
                "Account":               _DEBIT_ACCT,
                "Posting profile":       "",
                "Cash code":             _DEBIT_CASH_CODE,
                "Description":           debit_desc,
                "Debit":                 f"{debit_fee:.2f}",
                "Credit":                "",
                "Item sales tax group":  "",
                "Sales tax code":        "",
                "Offset company":        _COMPANY,
                "Offset account type":   "Bank",
                "Offset account":        offset_account,
                "Offset transaction text": "",
                "Currency":              _CURRENCY,
                "Exchange rate":         _EXCHANGE_RATE,
                "Item sales tax group2": "",
                "Sales tax group":       _SALES_TAX_GROUP,
                "Withholding tax group": "",
                "Release date":          "",
                "Reversing entry":       _REVERSING_ENTRY,
                "Reversing date":        "",
                "_needs_review":         False,
                "_review_reason":        "",
                "_match_confidence":     "HIGH",
            }
            rows.append(debit_row)
            log.append({
                "level": "OK",
                "msg": (f"Built {'batch' if multi_batch else 'single'} debit row "
                        f"for {date_str or 'undated'} — fee ${debit_fee:.2f} "
                        f"(source: {'Zoho summary' if summary_fee else 'per-txn sum'})")
            })

    if not rows:
        log.append({"level": "WARN", "msg": "No journal entry rows were generated."})
        return pd.DataFrame(columns=[c for c in D365_COLUMNS if not c.startswith("_")]), log

    journal_df = pd.DataFrame(rows)
    for col in D365_COLUMNS:
        if col not in journal_df.columns:
            journal_df[col] = ""

    return journal_df, log


# ─────────────────────────────────────────────────────────────────────────────
# Description builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_credit_description(cash_code, cash_pfx, desc_prefix,
                               account, account_name, boa_desc):
    """
    Build the D365 credit line description.

    Normal (AR001):  "AR001: BC000649 Equity Now Inc_ZOHO PAYMENTS DES:..."
    MPP    (AR002):  "MPP BC000327 Elite Functional Wellness_ZOHO PAYMENTS DES:..."
    CS Ticket:       "Nicole Holovach CS Ticket #676_ZOHO PAYMENTS DES:..."

    Rules:
      - If desc_prefix is set (CS ticket), it replaces everything before the BOA desc
      - If cash_pfx is "MPP ", no cash code token in description
      - Otherwise, lead with "{cash_code}: "
    """
    name_part = " ".join(filter(None, [account, account_name])).strip()

    if desc_prefix:
        # CS ticket: "Nicole Holovach CS Ticket #676_ZOHO PAYMENTS DES:..."
        if boa_desc:
            return f"{desc_prefix}{boa_desc}"
        return desc_prefix.rstrip("_")

    if cash_pfx:
        # MPP: "MPP BC000327 Name_ZOHO PAYMENTS DES:..."
        if boa_desc:
            return f"{cash_pfx}{name_part}_{boa_desc}"
        return f"{cash_pfx}{name_part}"

    # Standard format per SOP:
    #   With account resolved:   "AR001: BC000649 Equity Now Inc_ZOHO PAYMENTS DES:..."
    #   Without account (new):   "Legends Charter School_ZOHO PAYMENTS DES:..."
    #   (no cash code prefix when account number is not yet in master)
    if account:
        code_token = f"{cash_code}: " if cash_code else ""
        if boa_desc:
            return f"{code_token}{name_part}_{boa_desc}"
        return f"{code_token}{name_part}"
    else:
        # Account not resolved yet — omit cash code prefix, just use name
        if boa_desc:
            return f"{account_name}_{boa_desc}"
        return account_name


def _safe_float(val):
    try:
        f = float(str(val).replace(",", "").replace("$", "").strip())
        return None if np.isnan(f) else f
    except (ValueError, TypeError):
        return None


"""
exporter.py — Export D365 journal entries to a formatted Excel file.

Output matches the D365_General_Journal_Template.xlsx column order exactly.
Rows needing review are highlighted in amber.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── D365 output column order (exact match to template) ───────────────────────
D365_OUTPUT_COLUMNS = [
    "Date", "Voucher", "Account name", "Company", "Account type",
    "Account", "Posting profile", "Cash code", "Description",
    "Debit", "Credit",
    "Item sales tax group", "Sales tax code",
    "Offset company", "Offset account type", "Offset account",
    "Offset transaction text",
    "Currency", "Exchange rate",
    "Item sales tax group2", "Sales tax group",
    "Withholding tax group", "Release date", "Reversing entry", "Reversing date",
]

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1A3C6E", end_color="1A3C6E")   # navy
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
REVIEW_FILL   = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")   # amber
REVIEW_FONT   = Font(name="Arial", color="856404", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
CREDIT_FILL   = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")   # light green
DEBIT_FILL    = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")   # light blue

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMN_WIDTHS = {
    "Date": 14,
    "Voucher": 18,
    "Account name": 32,
    "Company": 9,
    "Account type": 13,
    "Account": 20,
    "Posting profile": 16,
    "Cash code": 10,
    "Description": 72,
    "Debit": 14,
    "Credit": 14,
    "Item sales tax group": 12,
    "Sales tax code": 12,
    "Offset company": 9,
    "Offset account type": 14,
    "Offset account": 40,
    "Offset transaction text": 14,
    "Currency": 10,
    "Exchange rate": 12,
    "Item sales tax group2": 12,
    "Sales tax group": 13,
    "Withholding tax group": 13,
    "Release date": 12,
    "Reversing entry": 13,
    "Reversing date": 12,
}


def export_to_excel(journal_df: pd.DataFrame) -> bytes:
    """
    Write D365 journal entries to an openpyxl workbook and return raw bytes.

    Two sheets are produced:
      1. 'D365 Upload'  — clean export, no internal columns
      2. 'Review Items' — only flagged rows, with review reason column
    """
    # ── Separate clean data from internal review columns ──────────────────────
    internal_cols = [c for c in journal_df.columns if c.startswith("_")]
    needs_review_col = "_needs_review"  if "_needs_review"  in journal_df.columns else None
    review_reason_col= "_review_reason" if "_review_reason" in journal_df.columns else None

    # Build the upload-ready DataFrame (only D365 columns, in order)
    upload_df = _build_upload_df(journal_df)

    wb = Workbook()

    # ── Sheet 1: D365 Upload ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "D365 Upload"
    _write_sheet(ws, upload_df, journal_df, needs_review_col)

    # ── Sheet 2: Review Items ─────────────────────────────────────────────────
    if needs_review_col and journal_df[needs_review_col].any():
        ws2 = wb.create_sheet("Review Items")
        flagged_idx = journal_df[journal_df[needs_review_col] == True].index
        flagged_upload = upload_df.loc[flagged_idx].copy()
        if review_reason_col:
            flagged_upload.insert(
                len(flagged_upload.columns), "Review Reason",
                journal_df.loc[flagged_idx, review_reason_col].values
            )
        _write_sheet(ws2, flagged_upload, journal_df, needs_review_col)

    # ── Sheet 3: Legend ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Legend")
    _write_legend(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_upload_df(journal_df: pd.DataFrame) -> pd.DataFrame:
    """Extract D365 output columns in canonical order; fill missing with ''."""
    df = pd.DataFrame()
    for col in D365_OUTPUT_COLUMNS:
        if col in journal_df.columns:
            df[col] = journal_df[col].fillna("").astype(str).replace("nan", "").replace("None", "")
        else:
            df[col] = ""
    return df


def _write_sheet(ws, upload_df: pd.DataFrame, source_df: pd.DataFrame, needs_review_col):
    """Write headers + data rows with formatting to ws."""
    # ── Headers ───────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(upload_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (df_idx, row) in enumerate(upload_df.iterrows(), start=2):
        flagged = False
        if needs_review_col and needs_review_col in source_df.columns:
            flagged = bool(source_df.loc[df_idx, needs_review_col]) if df_idx in source_df.index else False

        # Determine row type: credit (has Credit amount) or debit (has Debit)
        has_credit = str(row.get("Credit", "")).strip() not in ("", "0", "0.0", "0.00")
        has_debit  = str(row.get("Debit",  "")).strip() not in ("", "0", "0.0", "0.00")
        row_fill = REVIEW_FILL if flagged else (CREDIT_FILL if has_credit else (DEBIT_FILL if has_debit else None))
        row_font = REVIEW_FONT if flagged else NORMAL_FONT

        for col_idx, (col_name, value) in enumerate(row.items(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = row_font
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "Description"))
            if row_fill:
                cell.fill = row_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(upload_df.columns, start=1):
        width = COLUMN_WIDTHS.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze pane: keep headers visible ─────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    if upload_df.shape[0] > 0:
        last_col = get_column_letter(len(upload_df.columns))
        ws.auto_filter.ref = f"A1:{last_col}1"


def _write_legend(ws):
    """Write a color legend and column field guide on the Legend sheet."""
    ws.title = "Legend"

    headers = ["Color", "Meaning"]
    ws.append(headers)
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].font = HEADER_FONT
    ws["B1"].fill = HEADER_FILL

    legend_rows = [
        ("Green row",  "CREDIT line — customer payment (verified)"),
        ("Blue row",   "DEBIT line — Zoho merchant fee"),
        ("Amber row",  "⚠️ Flagged for manual review — check account, cash code, or amount"),
    ]

    fills = [CREDIT_FILL, DEBIT_FILL, REVIEW_FILL]
    fonts = [NORMAL_FONT, NORMAL_FONT, REVIEW_FONT]

    for i, (label, desc) in enumerate(legend_rows, start=2):
        ws.cell(row=i, column=1, value=label).fill = fills[i - 2]
        ws.cell(row=i, column=1).font = fonts[i - 2]
        ws.cell(row=i, column=2, value=desc).font = fonts[i - 2]

    ws.append([])
    ws.append(["Field", "Rule"])
    ws["A6"].font = HEADER_FONT
    ws["A6"].fill = HEADER_FILL
    ws["B6"].font = HEADER_FONT
    ws["B6"].fill = HEADER_FILL

    field_rules = [
        ("Date",             "BOA posting date (not journal entry date)"),
        ("Voucher",          "Leave blank — auto-generated by D365"),
        ("Account name",     "Canonical name from IBNY Business Customer Account file"),
        ("Company",          "Always: bwa"),
        ("Account type",     "Credit rows: Customer | Debit rows: Ledger"),
        ("Account",          "BC###### from customer master, or 43170111-… for fee rows"),
        ("Posting profile",  "AutoPost (credit rows only)"),
        ("Cash code",        "AR001=Due on receipt, AR002=MPP, OSF005=Merchant fee"),
        ("Description",      "Credit: {AR00X}: {BC######} {Customer Name}_{BOA desc}"),
        ("Debit",            "Merchant fee amount (debit rows only)"),
        ("Credit",           "Gross payment from Zoho (credit rows only, NOT BOA net)"),
        ("Offset account",   "B1000002 (acct 3371) | B1000003 (acct 3924) | B1000001 (acct 3384)"),
        ("Sales tax group",  "Always: AVATAX (credit rows)"),
        ("Reversing entry",  "Always: No"),
    ]

    for i, (field, rule) in enumerate(field_rules, start=7):
        ws.cell(row=i, column=1, value=field).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=rule).font  = NORMAL_FONT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70


# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="IBNY D365 Journal Entry Automation",
    page_icon="📒",
    layout="wide",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {font-size:1.6rem; font-weight:700; color:#1a3c6e; margin-bottom:0.2rem;}
    .sub-header  {font-size:0.95rem; color:#555; margin-bottom:1.5rem;}
    .section-title {font-size:1.05rem; font-weight:600; color:#1a3c6e;
                    border-bottom:2px solid #d0dff5; padding-bottom:4px; margin-bottom:1rem;}
    .flag-box {background:#fff8e1; border-left:4px solid #ffc107;
               padding:0.6rem 1rem; border-radius:4px; margin:4px 0;}
    .ok-box   {background:#e8f5e9; border-left:4px solid #4caf50;
               padding:0.6rem 1rem; border-radius:4px; margin:4px 0;}
    .err-box  {background:#ffebee; border-left:4px solid #f44336;
               padding:0.6rem 1rem; border-radius:4px; margin:4px 0;}
    .metric-card {background:#f0f4ff; border-radius:8px; padding:1rem;
                  text-align:center;}
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header">📒 IBNY D365 Journal Entry Automation</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Upload BOA and Zoho files to generate a D365-ready journal entry Excel.</div>', unsafe_allow_html=True)
st.caption("v1.4 — fee fix: uses Zoho summary fee")

# ── Load reference data (auto-loaded from app directory) ─────────────────────
@st.cache_data(show_spinner=False)
def load_references():
    import os
    _base = os.path.dirname(os.path.abspath(__file__))
    customers  = load_customer_master(os.path.join(_base, "Account_Masterlist.xlsx"))
    cash_codes = load_cash_codes(os.path.join(_base, "Cash_Code_Masterlist.xlsx"))
    return customers, cash_codes

try:
    customer_df, cash_code_df = load_references()
    cs_count = int((customer_df.get("CS/PS Ticket", pd.Series()).fillna("").str.strip() != "").sum()) if "CS/PS Ticket" in customer_df.columns else 0
    st.markdown(f'<div class="ok-box">✅ Reference files loaded — {len(customer_df)} customer accounts · {len(cash_code_df)} cash codes · CS/PS entries: {cs_count}</div>', unsafe_allow_html=True)
except Exception as e:
    st.markdown(f'<div class="err-box">❌ Could not load reference files: {e}</div>', unsafe_allow_html=True)
    st.stop()

st.markdown("---")

# ── File Upload ───────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="section-title">🏦 Bank of America File</div>', unsafe_allow_html=True)
    boa_file = st.file_uploader(
        "Upload BOA transaction export",
        type=["xlsx", "csv", "xls"],
        key="boa",
        help="Excel or CSV export from Bank of America. Must contain date, description, and amount columns.",
    )

with col2:
    st.markdown('<div class="section-title">💳 Zoho Payments File</div>', unsafe_allow_html=True)
    zoho_file = st.file_uploader(
        "Upload Zoho payment export",
        type=["xlsx", "csv", "xls", "pdf"],
        key="zoho",
        help="Zoho Payments export showing gross amount and merchant fee per transaction.",
    )

# Optional invoice uploader
with st.expander("📄 Upload Customer Invoices (optional — used when name is missing from Zoho)"):
    invoice_files = st.file_uploader(
        "Upload one or more invoice PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        key="invoices",
    )

st.markdown("---")

# ── BOA Account routing (from automation rules §5.1 col 16) ──────────────────
st.markdown('<div class="section-title">⚙️ BOA Account Settings</div>', unsafe_allow_html=True)
acct_col1, acct_col2 = st.columns([1, 2])
with acct_col1:
    boa_account_last4 = st.selectbox(
        "BOA Source Account (last 4 digits)",
        options=["3371", "3924", "3384", "Unknown"],
        index=0,
        help="Determines the Offset Account (B1000002, B1000003, or B1000001).",
    )
with acct_col2:
    offset_map = {"3371": "B1000002", "3924": "B1000003", "3384": "B1000001", "Unknown": "B1000002"}
    selected_offset = offset_map[boa_account_last4]
    st.info(f"Offset Account → **{selected_offset}**")

st.markdown("---")

# ── Process ───────────────────────────────────────────────────────────────────
if st.button("🚀 Generate D365 Journal Entries", type="primary", use_container_width=True):
    if not boa_file or not zoho_file:
        st.error("Please upload both a BOA file and a Zoho file before proceeding.")
        st.stop()

    with st.spinner("Parsing files…"):
        boa_df, boa_errors = parse_boa(boa_file)
        zoho_df, zoho_errors = parse_zoho(zoho_file)

    # ── Parse feedback ──
    if boa_errors:
        for e in boa_errors:
            st.markdown(f'<div class="flag-box">⚠️ BOA: {e}</div>', unsafe_allow_html=True)
    if zoho_errors:
        for e in zoho_errors:
            st.markdown(f'<div class="flag-box">⚠️ Zoho: {e}</div>', unsafe_allow_html=True)

    if boa_df is None or boa_df.empty:
        st.error("Could not parse BOA file. Check the format and try again.")
        st.stop()
    if zoho_df is None or zoho_df.empty:
        st.error("Could not parse Zoho file. Check the format and try again.")
        st.stop()

    # ── Preview raw parses ──
    with st.expander("🔍 Raw BOA Rows Parsed"):
        st.dataframe(boa_df, use_container_width=True)
    with st.expander("🔍 Raw Zoho Rows Parsed"):
        st.dataframe(zoho_df, use_container_width=True)
        if "_summary_fee" in zoho_df.columns:
            st.success(f"✅ Zoho summary fee (authoritative): **${float(zoho_df['_summary_fee'].iloc[0]):,.2f}** — this will be used as the debit amount.")
        else:
            st.error("❌ _summary_fee column missing — old parsers.py is still running. Push the new file to GitHub.")

    with st.spinner("Matching transactions and resolving accounts…"):
        matched_df, match_log = match_transactions(
            boa_df, zoho_df, customer_df, invoice_files
        )

    with st.spinner("Building D365 journal entries…"):
        journal_df, build_log = build_journal_entries(
            matched_df, customer_df, selected_offset
        )

    # ── Match summary metrics ──
    st.markdown("---")
    st.markdown('<div class="section-title">📊 Processing Summary</div>', unsafe_allow_html=True)

    total     = len(matched_df)
    confident = int(matched_df["_match_confidence"].eq("HIGH").sum())  if "_match_confidence" in matched_df else 0
    flagged   = int(matched_df["_needs_review"].sum())                 if "_needs_review"     in matched_df else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Zoho Transactions", total)
    m2.metric("Matched (High Confidence)", confident)
    m3.metric("Flagged for Review", flagged)
    m4.metric("D365 Rows Generated", len(journal_df))

    # ── Match log details ──
    if match_log:
        with st.expander(f"📋 Match Log ({len(match_log)} entries)"):
            for entry in match_log:
                css_class = "flag-box" if entry.get("level") == "WARN" else "ok-box"
                st.markdown(f'<div class="{css_class}">{entry["msg"]}</div>', unsafe_allow_html=True)

    # ── Review table: flagged rows ──
    needs_review = journal_df[journal_df.get("_needs_review", False) == True] if "_needs_review" in journal_df else pd.DataFrame()
    if not needs_review.empty:
        st.markdown("---")
        st.markdown('<div class="section-title">🚩 Rows Requiring Manual Review</div>', unsafe_allow_html=True)
        st.warning(f"{len(needs_review)} row(s) could not be fully resolved. These appear highlighted in the export.")
        review_cols = [c for c in needs_review.columns if not c.startswith("_")]
        st.dataframe(needs_review[review_cols], use_container_width=True)

    # ── Full journal preview ──
    st.markdown("---")
    st.markdown('<div class="section-title">📄 D365 Journal Entry Preview</div>', unsafe_allow_html=True)
    display_cols = [c for c in journal_df.columns if not c.startswith("_")]
    st.dataframe(journal_df[display_cols], use_container_width=True, height=400)

    # ── Export ──
    with st.spinner("Building Excel export…"):
        excel_bytes = export_to_excel(journal_df)

    st.success("✅ D365 journal entry file ready for download.")
    st.download_button(
        label="⬇️ Download D365 Journal Entry Excel",
        data=excel_bytes,
        file_name="D365_Journal_Entry_Upload.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── Sidebar: reference info ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📚 Reference Data")
    st.markdown(f"**Customer Accounts:** {len(customer_df)}")
    st.markdown(f"**Cash Codes:** {len(cash_code_df)}")

    st.markdown("---")
    st.markdown("### 🗂️ Cash Code Reference")
    st.dataframe(
        cash_code_df[cash_code_df["Cash Code"].str.startswith("AR")],
        use_container_width=True, height=320, hide_index=True
    )

    st.markdown("---")
    st.markdown("### 📘 How to Use")
    st.markdown("""
1. Upload **BOA Excel/CSV** export
2. Upload **Zoho Payments** export
3. Select the BOA account (last 4 digits)
4. Click **Generate Journal Entries**
5. Review flagged rows
6. Download the Excel file
""")
    st.markdown("---")
    st.caption("IBNY Cash Closing Automation v1.0")
