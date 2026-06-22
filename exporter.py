"""
exporter.py — Export D365 journal entries to a formatted Excel file.

Output matches the D365_General_Journal_Template.xlsx column order exactly.
Rows needing review are highlighted in amber.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── D365 output column order (exact match to template) ───────────────────────
D365_OUTPUT_COLUMNS = [
    "Date", "Voucher", "Account name", "Company", "Account type",
    "Account", "Posting profile", "Cash code", "Description",
    "Debit", "Credit",
    "Item sales tax group", "Sales tax code",
    "Offset company", "Offset account type", "Offset account",
    "Offset transaction text",
    "Currency", "Exchange rate",
    "Item sales tax group2", "Sales tax group",
    "Withholding tax group", "Release date", "Reversing entry", "Reversing date",
]

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1A3C6E", end_color="1A3C6E")   # navy
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
REVIEW_FILL   = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")   # amber
REVIEW_FONT   = Font(name="Arial", color="856404", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
CREDIT_FILL   = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")   # light green
DEBIT_FILL    = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")   # light blue

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMN_WIDTHS = {
    "Date": 14,
    "Voucher": 18,
    "Account name": 32,
    "Company": 9,
    "Account type": 13,
    "Account": 20,
    "Posting profile": 16,
    "Cash code": 10,
    "Description": 72,
    "Debit": 14,
    "Credit": 14,
    "Item sales tax group": 12,
    "Sales tax code": 12,
    "Offset company": 9,
    "Offset account type": 14,
    "Offset account": 40,
    "Offset transaction text": 14,
    "Currency": 10,
    "Exchange rate": 12,
    "Item sales tax group2": 12,
    "Sales tax group": 13,
    "Withholding tax group": 13,
    "Release date": 12,
    "Reversing entry": 13,
    "Reversing date": 12,
}


def export_to_excel(journal_df: pd.DataFrame) -> bytes:
    """
    Write D365 journal entries to an openpyxl workbook and return raw bytes.

    Two sheets are produced:
      1. 'D365 Upload'  — clean export, no internal columns
      2. 'Review Items' — only flagged rows, with review reason column
    """
    # ── Separate clean data from internal review columns ──────────────────────
    internal_cols = [c for c in journal_df.columns if c.startswith("_")]
    needs_review_col = "_needs_review"  if "_needs_review"  in journal_df.columns else None
    review_reason_col= "_review_reason" if "_review_reason" in journal_df.columns else None

    # Build the upload-ready DataFrame (only D365 columns, in order)
    upload_df = _build_upload_df(journal_df)

    wb = Workbook()

    # ── Sheet 1: D365 Upload ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "D365 Upload"
    _write_sheet(ws, upload_df, journal_df, needs_review_col)

    # ── Sheet 2: Review Items ─────────────────────────────────────────────────
    if needs_review_col and journal_df[needs_review_col].any():
        ws2 = wb.create_sheet("Review Items")
        flagged_idx = journal_df[journal_df[needs_review_col] == True].index
        flagged_upload = upload_df.loc[flagged_idx].copy()
        if review_reason_col:
            flagged_upload.insert(
                len(flagged_upload.columns), "Review Reason",
                journal_df.loc[flagged_idx, review_reason_col].values
            )
        _write_sheet(ws2, flagged_upload, journal_df, needs_review_col)

    # ── Sheet 3: Legend ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Legend")
    _write_legend(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_upload_df(journal_df: pd.DataFrame) -> pd.DataFrame:
    """Extract D365 output columns in canonical order; fill missing with ''."""
    df = pd.DataFrame()
    for col in D365_OUTPUT_COLUMNS:
        if col in journal_df.columns:
            df[col] = journal_df[col].fillna("").astype(str).replace("nan", "").replace("None", "")
        else:
            df[col] = ""
    return df


def _write_sheet(ws, upload_df: pd.DataFrame, source_df: pd.DataFrame, needs_review_col):
    """Write headers + data rows with formatting to ws."""
    # ── Headers ───────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(upload_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (df_idx, row) in enumerate(upload_df.iterrows(), start=2):
        flagged = False
        if needs_review_col and needs_review_col in source_df.columns:
            flagged = bool(source_df.loc[df_idx, needs_review_col]) if df_idx in source_df.index else False

        # Determine row type: credit (has Credit amount) or debit (has Debit)
        has_credit = str(row.get("Credit", "")).strip() not in ("", "0", "0.0", "0.00")
        has_debit  = str(row.get("Debit",  "")).strip() not in ("", "0", "0.0", "0.00")
        row_fill = REVIEW_FILL if flagged else (CREDIT_FILL if has_credit else (DEBIT_FILL if has_debit else None))
        row_font = REVIEW_FONT if flagged else NORMAL_FONT

        for col_idx, (col_name, value) in enumerate(row.items(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = row_font
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "Description"))
            if row_fill:
                cell.fill = row_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(upload_df.columns, start=1):
        width = COLUMN_WIDTHS.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze pane: keep headers visible ─────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    if upload_df.shape[0] > 0:
        last_col = get_column_letter(len(upload_df.columns))
        ws.auto_filter.ref = f"A1:{last_col}1"


def _write_legend(ws):
    """Write a color legend and column field guide on the Legend sheet."""
    ws.title = "Legend"

    headers = ["Color", "Meaning"]
    ws.append(headers)
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].font = HEADER_FONT
    ws["B1"].fill = HEADER_FILL

    legend_rows = [
        ("Green row",  "CREDIT line — customer payment (verified)"),
        ("Blue row",   "DEBIT line — Zoho merchant fee"),
        ("Amber row",  "⚠️ Flagged for manual review — check account, cash code, or amount"),
    ]

    fills = [CREDIT_FILL, DEBIT_FILL, REVIEW_FILL]
    fonts = [NORMAL_FONT, NORMAL_FONT, REVIEW_FONT]

    for i, (label, desc) in enumerate(legend_rows, start=2):
        ws.cell(row=i, column=1, value=label).fill = fills[i - 2]
        ws.cell(row=i, column=1).font = fonts[i - 2]
        ws.cell(row=i, column=2, value=desc).font = fonts[i - 2]

    ws.append([])
    ws.append(["Field", "Rule"])
    ws["A6"].font = HEADER_FONT
    ws["A6"].fill = HEADER_FILL
    ws["B6"].font = HEADER_FONT
    ws["B6"].fill = HEADER_FILL

    field_rules = [
        ("Date",             "BOA posting date (not journal entry date)"),
        ("Voucher",          "Leave blank — auto-generated by D365"),
        ("Account name",     "Canonical name from IBNY Business Customer Account file"),
        ("Company",          "Always: bwa"),
        ("Account type",     "Credit rows: Customer | Debit rows: Ledger"),
        ("Account",          "BC###### from customer master, or 43170111-… for fee rows"),
        ("Posting profile",  "AutoPost (credit rows only)"),
        ("Cash code",        "AR001=Due on receipt, AR002=MPP, OSF005=Merchant fee"),
        ("Description",      "Credit: {AR00X}: {BC######} {Customer Name}_{BOA desc}"),
        ("Debit",            "Merchant fee amount (debit rows only)"),
        ("Credit",           "Gross payment from Zoho (credit rows only, NOT BOA net)"),
        ("Offset account",   "B1000002 (acct 3371) | B1000003 (acct 3924) | B1000001 (acct 3384)"),
        ("Sales tax group",  "Always: AVATAX (credit rows)"),
        ("Reversing entry",  "Always: No"),
    ]

    for i, (field, rule) in enumerate(field_rules, start=7):
        ws.cell(row=i, column=1, value=field).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=rule).font  = NORMAL_FONT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70
